import openpyxl, json, sys, re, pathlib
from collections import defaultdict

def extrair(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    def v(ws, cell):
        val = ws[cell].value
        return val if val is not None else 0

    PT_EN = {
        'BRASIL':'Brazil','MÉXICO':'Mexico','MEXICO':'Mexico','CHILE':'Chile','ARGENTINA':'Argentina',
        'COLÔMBIA':'Colombia','COLOMBIA':'Colombia','PERU':'Peru','URUGUAI':'Uruguay','PARAGUAI':'Paraguay',
        'BOLÍVIA':'Bolivia','BOLIVIA':'Bolivia','EQUADOR':'Ecuador','VENEZUELA':'Venezuela','CUBA':'Cuba',
        'ESTADOS UNIDOS':'United States of America','EUA':'United States of America','CANADÁ':'Canada','CANADA':'Canada',
        'PORTUGAL':'Portugal','ESPANHA':'Spain','FRANÇA':'France','FRANCA':'France','ITÁLIA':'Italy','ITALIA':'Italy',
        'ALEMANHA':'Germany','REINO UNIDO':'United Kingdom','SUÍÇA':'Switzerland','SUICA':'Switzerland',
        'MOÇAMBIQUE':'Mozambique','MOCAMBIQUE':'Mozambique','ANGOLA':'Angola','CABO VERDE':'Cabo Verde',
        'GUINÉ-BISSAU':'Guinea-Bissau','SÃO TOMÉ E PRÍNCIPE':'São Tomé and Principe','TIMOR-LESTE':'Timor-Leste',
        'CHINA':'China', 'JAPÃO':'Japan', 'JAPAO':'Japan', 'ÍNDIA':'India', 'INDIA':'India',
        'ÁFRICA DO SUL':'South Africa', 'AFRICA DO SUL':'South Africa',
    }
    def norm_pais(pais):
        if not pais: return None
        p = str(pais).strip().upper()
        return PT_EN.get(p, p.title())

    def pct_status(pct):
        if pct <= 0: return 'NÃO INICIADO'
        if pct >= 1: return 'CONCLUÍDO'
        return 'EM ANDAMENTO'

    result = {}

    # ================= RESUMO =================
    ws = wb['📊 Resumo']
    result['resumo'] = {
        'total_rede': {'aprovado': v(ws,'B6'), 'executado': v(ws,'B7'), 'saldo': v(ws,'B8')},
        'bolsas_cg': {'previsto': v(ws,'D6'), 'executado': v(ws,'D7'), 'saldo': v(ws,'D8')},
        'missoes_cg': {'aprovado': v(ws,'F6'), 'executado': v(ws,'F7'), 'saldo': v(ws,'F8')},
        'acoes_inst': {'aprovado': v(ws,'H6'), 'executado': v(ws,'H7'), 'saldo': v(ws,'H8')},
        'missoes_temas': {'aprovado': v(ws,'J6'), 'executado': v(ws,'J7'), 'saldo': v(ws,'J8')},
        'bolsas_temas': {'planejado': v(ws,'L6'), 'executado': v(ws,'L7'), 'saldo': v(ws,'L8')},
    }
    anos = []
    for r in range(12, 17):
        anos.append({'ano': ws[f'A{r}'].value, 'previsto': v(ws,f'B{r}'), 'executado': v(ws,f'C{r}'), 'saldo': v(ws,f'D{r}'), 'pct': v(ws,f'E{r}')})
    result['resumo']['por_ano'] = anos

    # ================= GERAL =================
    ws2 = wb['📈 Geral']
    geral = []
    for row in ws2.iter_rows(min_row=5, max_row=ws2.max_row, values_only=True):
        if row[0] and row[1] and isinstance(row[0], str):
            if row[0].startswith('Total') or row[0].startswith('🌎') or 'Executado por Ano' in str(row[0]):
                continue
            geral.append({'ies': row[0], 'modalidade': row[1], '2026': row[2] or 0, '2027': row[3] or 0,
                           '2028': row[4] or 0, '2029': row[5] or 0, '2030': row[6] or 0,
                           'total_previsto': row[7] or 0, 'total_executado': row[8] or 0})
    result['geral_matrix'] = geral

    # ================= CG =================
    wscg = wb['💼 CG']
    por_ies_cg = []
    for r in range(6, 12):
        por_ies_cg.append({'ies': wscg.cell(r,1).value, 'missoes_prev': v(wscg,f'B{r}'), 'missoes_exec': v(wscg,f'C{r}'),
            'acoes_prev': v(wscg,f'D{r}'), 'acoes_exec': v(wscg,f'E{r}'), 'bolsas_prev': v(wscg,f'F{r}'), 'bolsas_exec': v(wscg,f'G{r}'),
            'previsto': v(wscg,f'H{r}'), 'executado': v(wscg,f'I{r}'), 'saldo': v(wscg,f'J{r}'), 'pct': v(wscg,f'K{r}')})
    totals_row = {'previsto': v(wscg,'H12'), 'executado': v(wscg,'I12'), 'saldo': v(wscg,'J12'), 'pct': v(wscg,'K12')}
    por_modalidade_cg = []
    for r in range(16, 19):
        por_modalidade_cg.append({'modalidade': wscg.cell(r,1).value, 'previsto': v(wscg,f'B{r}'), 'executado': v(wscg,f'C{r}'), 'saldo': v(wscg,f'D{r}'), 'pct': v(wscg,f'E{r}')})
    por_ano_cg = []
    for r in range(23, 28):
        por_ano_cg.append({'ano': wscg.cell(r,1).value, 'previsto': v(wscg,f'B{r}'), 'executado': v(wscg,f'C{r}'), 'saldo': v(wscg,f'D{r}'), 'pct': v(wscg,f'E{r}')})
    result['cg'] = {'kpis': {**totals_row, 'status': pct_status(totals_row['pct'])}, 'por_ies': por_ies_cg,
                     'por_modalidade': por_modalidade_cg, 'por_ano': por_ano_cg}

    # ================= INSTITUIÇÕES =================
    inst_sheets = {'Fiocruz': '🏰Fiocruz', 'UFMS': '🏛️UFMS', 'UNILA': '🏛️UNILA', 'UNIR': '🏛️UNIR', 'UFOPA': '🏛️UFOPA', 'UFPI': '🏛️UFPI'}
    instituicoes = {}
    for nome, sheet_name in inst_sheets.items():
        wsi = wb[sheet_name]
        total_prev = v(wsi,'H4'); total_exec = v(wsi,'I4'); total_saldo = v(wsi,'J4')
        total_pct = total_exec / total_prev if total_prev else 0
        modalidades = []
        for r in range(7, 11):
            mod = wsi.cell(r,1).value
            if not mod: continue
            modalidades.append({'modalidade': mod, '2026': v(wsi,f'B{r}'), '2027': v(wsi,f'C{r}'), '2028': v(wsi,f'D{r}'),
                                 '2029': v(wsi,f'E{r}'), '2030': v(wsi,f'F{r}'), 'total_previsto': v(wsi,f'G{r}'),
                                 'total_executado': v(wsi,f'H{r}'), 'saldo': v(wsi,f'I{r}'), 'pct': v(wsi,f'J{r}')})
        anos_i = []
        for r in range(17, 22):
            ano_val = wsi.cell(r,1).value
            if not ano_val: continue
            anos_i.append({'ano': ano_val, 'previsto': v(wsi,f'B{r}'), 'executado': v(wsi,f'C{r}'), 'saldo': v(wsi,f'D{r}'), 'pct': v(wsi,f'E{r}')})
        instituicoes[nome] = {'kpis': {'previsto': total_prev, 'executado': total_exec, 'saldo': total_saldo, 'pct': total_pct, 'status': pct_status(total_pct)},
                               'modalidades': modalidades, 'por_ano': anos_i}
    result['instituicoes'] = instituicoes

    # ================= TEMAS overview =================
    wst = wb['📚Temas']
    temas_overview_rows = []
    for row in wst.iter_rows(min_row=4, max_row=wst.max_row, values_only=True):
        if row[0] and row[1]:
            temas_overview_rows.append({'tema': row[0], 'componente': row[1], '2026': row[2] or 0, '2027': row[3] or 0,
                                         '2028': row[4] or 0, '2029': row[5] or 0, '2030': row[6] or 0,
                                         'total_previsto': row[7] or 0, 'total_executado': row[8] or 0})
    result['temas_overview'] = temas_overview_rows

    # ================= TEMAS individuais (com detalhamento por ano) =================
    tema_sheets = {'Tema 01': '🦠 Tema 01', 'Tema 02': '🚑Tema 02', 'Tema 03': '🌱Tema 03',
                   'Tema 04': '👨\u200d👩\u200d👧\u200d👦 Tema 04', 'Tema 05': '💡Tema 05'}
    temas = {}
    for nome, sheet_name in tema_sheets.items():
        wst2 = wb[sheet_name]
        total_prev = v(wst2,'F4'); total_exec = v(wst2,'G4'); total_saldo = v(wst2,'H4')
        total_pct = total_exec / total_prev if total_prev else 0

        missoes_ies = []
        for r in range(8, 14):
            ies = wst2.cell(r,1).value
            if not ies: continue
            missoes_ies.append({'ies': ies, 'total_previsto': v(wst2,f'G{r}'), 'total_executado': v(wst2,f'H{r}')})

        # Previsto por ano = soma das missões por IES (linhas 8-13, colunas B-F) + bolsas do tema (linha 16, colunas B-F)
        anos5 = ['2026','2027','2028','2029','2030']
        cols = ['B','C','D','E','F']
        previsto_por_ano = {a: 0 for a in anos5}
        for r in range(8, 14):
            for a, c in zip(anos5, cols):
                previsto_por_ano[a] += v(wst2, f'{c}{r}')
        bolsas_row = 16  # 'TEMA XX - Bolsas' linha logo após subtotal missões
        # localizar dinamicamente a linha "Bolsas do Tema" (rótulo variável)
        for r in range(14, 18):
            label = wst2.cell(r,1).value
            if label and 'Bolsas' in str(label):
                bolsas_row = r
                break
        for a, c in zip(anos5, cols):
            previsto_por_ano[a] += v(wst2, f'{c}{bolsas_row}')

        # Executado por ano: linha TOTAL da tabela "EXECUÇÃO POR ANO - RUBRICA"
        executado_por_ano = {a: 0 for a in anos5}
        exec_total_row = None
        for r in range(1, wst2.max_row+1):
            if wst2.cell(r,1).value == 'TOTAL' and wst2.cell(r-1,1).value and 'Bolsas' in str(wst2.cell(r-1,1).value):
                exec_total_row = r
                break
        if exec_total_row:
            for a, c in zip(anos5, cols):
                executado_por_ano[a] = v(wst2, f'{c}{exec_total_row}')

        por_ano = []
        for a in anos5:
            p = previsto_por_ano[a]; e = executado_por_ano[a]
            por_ano.append({'ano': a, 'previsto': p, 'executado': e, 'saldo': p - e, 'pct': (e/p if p else 0)})

        temas[nome] = {'kpis': {'previsto': total_prev, 'executado': total_exec, 'saldo': total_saldo, 'pct': total_pct, 'status': pct_status(total_pct)},
                        'missoes_por_ies': missoes_ies, 'por_ano': por_ano}
    result['temas'] = temas

    # ================= BOLSAS (reestruturado) =================
    wsb = wb['🎒 Bolsas']

    # Quantidade de bolsas por IES / Tema / Modalidade (tabela "Quantidade de Bolsas", col O=15 em diante)
    qtd_header = [wsb.cell(5,c).value for c in range(15,28)]  # Modalidade, FIOCRUZ..UFPI, Tema01..05, TOTAL
    qtd_por_modalidade = []
    qtd_por_ies = defaultdict(int)
    qtd_por_tema = defaultdict(int)
    for r in range(6, 14):
        modal = wsb.cell(r,15).value
        if not modal: continue
        vals = {qtd_header[i]: (wsb.cell(r,15+i).value or 0) for i in range(1,13)}  # skip 'Modalidade' col itself
        qtd_por_modalidade.append({'modalidade': modal, **vals})
        for ies in ['FIOCRUZ','UNILA','UFMS','UNIR','UFOPA','UFPI']:
            qtd_por_ies[ies] += vals.get(ies, 0)
        for tema in ['Tema 01','Tema 02','Tema 03','Tema 04','Tema 05']:
            qtd_por_tema[tema] += vals.get(tema, 0)

    # Totais por modalidade (valor executado + qtd ativa/encerrada) — colunas Z(26)..AD(30)
    modalidade_totais = []
    for r in range(20, 28):
        mod = wsb.cell(r,26).value
        if not mod: continue
        modalidade_totais.append({'modalidade': mod, 'qtd_total': wsb.cell(r,27).value or 0,
            'qtd_ativa': wsb.cell(r,28).value or 0, 'qtd_encerrada': wsb.cell(r,29).value or 0,
            'valor_executado': wsb.cell(r,30).value or 0})

    # Publicações / artigos científicos por modalidade — colunas AC(29)/AD(30)
    publicacoes_por_modalidade = []
    for r in range(6, 14):
        mod = wsb.cell(r,29).value
        if not mod: continue
        publicacoes_por_modalidade.append({'modalidade': mod, 'qtd_publicacoes': wsb.cell(r,30).value or 0})
    total_publicacoes = wsb.cell(15,30).value or 0

    # Registros individuais (usados só internamente para o mapa de calor por país/IES destino — não exibidos em tabela)
    registros_bolsas = []
    header_row = 29
    for r in range(header_row+1, wsb.max_row+1):
        nome = wsb.cell(r,2).value; modal = wsb.cell(r,4).value; pais = wsb.cell(r,7).value
        if not (nome or modal or pais): continue
        registros_bolsas.append({'modalidade': modal, 'ies_sede': wsb.cell(r,5).value, 'ies_destino': wsb.cell(r,6).value,
            'pais_destino': pais, 'tema': wsb.cell(r,8).value, 'valor_total': wsb.cell(r,12).value or 0, 'status': wsb.cell(r,13).value})

    result['bolsas'] = {
        'qtd_por_modalidade': qtd_por_modalidade,
        'qtd_por_ies': dict(qtd_por_ies),
        'qtd_por_tema': dict(qtd_por_tema),
        'modalidade_totais': modalidade_totais,
        'publicacoes_por_modalidade': publicacoes_por_modalidade,
        'total_publicacoes': total_publicacoes,
        'registros': registros_bolsas,
    }

    # ================= MISSÕES =================
    wsm = wb['✈️Missões']

    def read_ano_table(ws, search_start, search_end):
        header_row = None
        for r in range(search_start, search_end):
            if str(ws.cell(r,1).value).strip() == 'Ano':
                header_row = r
                break
        rows = []
        if header_row:
            r = header_row + 1
            while True:
                ano_val = ws.cell(r,1).value
                if not ano_val or str(ano_val).strip().upper() == 'TOTAL':
                    break
                rows.append({'ano': ano_val, 'programada': ws.cell(r,2).value or 0, 'realizada': ws.cell(r,3).value or 0,
                             'saldo': ws.cell(r,4).value or 0, 'pct': ws.cell(r,5).value or 0})
                r += 1
        return rows, (header_row or search_start)

    cg_por_ano, _end_cg = read_ano_table(wsm, 1, 20)
    temas_por_ano, _ = read_ano_table(wsm, _end_cg + 1, _end_cg + 20)
    registros_missoes = []
    for r in range(26, wsm.max_row+1):
        tipo = wsm.cell(r,2).value
        if not tipo: continue
        registros_missoes.append({'tipo': tipo, 'tema': wsm.cell(r,3).value, 'ies': wsm.cell(r,4).value, 'ano': wsm.cell(r,5).value,
            'destino': wsm.cell(r,6).value, 'participantes': wsm.cell(r,7).value, 'dias': wsm.cell(r,8).value,
            'valor_total': wsm.cell(r,9).value or 0, 'status': wsm.cell(r,10).value})
    result['missoes'] = {'cg_por_ano': cg_por_ano, 'temas_por_ano': temas_por_ano, 'registros': registros_missoes}

    # ================= EVENTOS (com resumo) =================
    wse = wb['👩\u200d🏫Eventos']
    eventos_resumo = {
        'por_situacao': [{'situacao': wse.cell(r,1).value, 'qtd': wse.cell(r,2).value or 0} for r in range(6,9) if wse.cell(r,1).value],
        'por_abrangencia': [{'abrangencia': wse.cell(r,4).value, 'qtd': wse.cell(r,5).value or 0} for r in range(6,10) if wse.cell(r,4).value],
        'por_modalidade': [{'modalidade': wse.cell(r,7).value, 'qtd': wse.cell(r,8).value or 0} for r in range(6,9) if wse.cell(r,7).value],
        'total': wse.cell(6,10).value or 0,
    }
    registros_eventos = []
    header_row_ev = 14
    for r in range(1, 25):
        if str(wse.cell(r,1).value).strip() == 'Nº':
            header_row_ev = r
            break
    for r in range(header_row_ev + 1, wse.max_row+1):
        titulo = wse.cell(r,3).value
        tipo_ev = wse.cell(r,4).value
        pais_ev = wse.cell(r,9).value
        if not (titulo or tipo_ev or pais_ev): continue
        registros_eventos.append({'ano': wse.cell(r,2).value, 'titulo': titulo or '—', 'tipo': tipo_ev,
            'abrangencia': wse.cell(r,5).value, 'finalidade': wse.cell(r,6).value, 'modalidade': wse.cell(r,7).value,
            'responsavel': wse.cell(r,8).value, 'pais': pais_ev, 'situacao': wse.cell(r,12).value})
    result['eventos'] = {'resumo': eventos_resumo, 'registros': registros_eventos}

    # ================= COMUNICAÇÃO =================
    wscm = wb['📢 Comunicação']
    pub_por_tipo = []
    header_row_com = None
    for r in range(1, 15):
        if str(wscm.cell(r,1).value).strip() == 'Tipo':
            header_row_com = r
            break
    if header_row_com:
        r = header_row_com + 1
        while True:
            tipo = wscm.cell(r,1).value
            if not tipo or str(tipo).strip().upper() == 'TOTAL':
                break
            pub_por_tipo.append({'tipo': tipo, 'qtd': wscm.cell(r,2).value or 0})
            r += 1
    visualizacoes_total = wscm.cell(header_row_com, 4).value if header_row_com else 0
    result['comunicacao'] = {'publicacoes_por_tipo': pub_por_tipo, 'visualizacoes_total': visualizacoes_total or 0}

    # ================= RI (Relações Internacionais) =================
    wri = wb['🌍 RI']
    ri_resumo = {
        'por_tipo_interacao': [{'tipo': wri.cell(r,1).value, 'qtd': wri.cell(r,2).value or 0} for r in range(6,9) if wri.cell(r,1).value],
        'por_tipo_acordo': [{'tipo': wri.cell(r,4).value, 'qtd': wri.cell(r,5).value or 0} for r in range(6,10) if wri.cell(r,4).value],
        'por_situacao': [{'situacao': wri.cell(r,7).value, 'qtd': wri.cell(r,8).value or 0} for r in range(6,10) if wri.cell(r,7).value],
        'total_registros': wri.cell(6,10).value or 0,
        'instituicoes_parceiras': wri.cell(9,10).value or 0,
        'paises_alcancados': wri.cell(12,10).value or 0,
        'por_continente': [{'continente': wri.cell(r,1).value, 'qtd': wri.cell(r,2).value or 0} for r in range(15,22) if wri.cell(r,1).value],
    }
    registros_ri = []
    for r in range(26, wri.max_row+1):
        inst = wri.cell(r,3).value
        if not inst: continue
        registros_ri.append({'data': wri.cell(r,2).value, 'instituicao': inst, 'pais': wri.cell(r,4).value,
            'continente': wri.cell(r,5).value, 'tipo_interacao': wri.cell(r,6).value, 'tipo_acordo': wri.cell(r,7).value,
            'ies_rede': wri.cell(r,8).value, 'objetivo': wri.cell(r,9).value, 'situacao': wri.cell(r,10).value})
    result['ri'] = {'resumo': ri_resumo, 'registros': registros_ri}

    # ================= PAÍSES (agregados p/ mapas de calor) =================
    def add_registros_paises(lst, registros, campo_pais, origem, campo_valor=None, campo_extra=None):
        for reg in registros:
            pais = reg.get(campo_pais)
            if pais:
                item = {'origem': origem, 'pais': str(pais).strip(), 'pais_en': norm_pais(pais),
                         'valor': reg.get(campo_valor, 0) if campo_valor else 0}
                if campo_extra:
                    item['extra'] = reg.get(campo_extra)
                lst.append(item)

    paises_bolsas = []
    add_registros_paises(paises_bolsas, registros_bolsas, 'pais_destino', 'Bolsas', 'valor_total', 'ies_destino')

    paises_missoes = []
    add_registros_paises(paises_missoes, registros_missoes, 'destino', 'Missões', 'valor_total')

    paises_eventos = []
    add_registros_paises(paises_eventos, registros_eventos, 'pais', 'Eventos')

    paises_ri = []
    add_registros_paises(paises_ri, registros_ri, 'pais', 'RI')

    def aggregate_paises(lst):
        agg = {}
        for reg in lst:
            key = reg['pais_en']
            if key not in agg:
                agg[key] = {'pais': reg['pais'], 'pais_en': key, 'total_registros': 0, 'valor_total': 0, 'por_origem': {}}
            agg[key]['total_registros'] += 1
            agg[key]['valor_total'] += reg['valor']
            agg[key]['por_origem'][reg['origem']] = agg[key]['por_origem'].get(reg['origem'], 0) + 1
        return list(agg.values())

    result['paises'] = {
        'bolsas': {'registros': paises_bolsas, 'por_pais': aggregate_paises(paises_bolsas)},
        'missoes': {'registros': paises_missoes, 'por_pais': aggregate_paises(paises_missoes)},
        'eventos': {'registros': paises_eventos, 'por_pais': aggregate_paises(paises_eventos)},
        'ri': {'registros': paises_ri, 'por_pais': aggregate_paises(paises_ri)},
        'geral': {'por_pais': aggregate_paises(paises_bolsas + paises_missoes + paises_eventos + paises_ri)},
    }

    # qtd_missoes derivadas (para os cards de ano)
    qtd_por_ies_ano = defaultdict(int)
    for reg in registros_missoes:
        if reg['ies'] and reg['ano']:
            qtd_por_ies_ano[(reg['ies'], str(reg['ano']))] += 1
    for nome, inst in instituicoes.items():
        for ano_obj in inst['por_ano']:
            ano_obj['qtd_missoes'] = qtd_por_ies_ano.get((nome, str(ano_obj['ano'])), 0)
    qtd_cg_ano = defaultdict(int)
    for reg in registros_missoes:
        if reg['tipo'] == 'Missões CG' and reg['ano']:
            qtd_cg_ano[str(reg['ano'])] += 1
    for ano_obj in result['cg']['por_ano']:
        ano_obj['qtd_missoes'] = qtd_cg_ano.get(str(ano_obj['ano']), 0)
    qtd_tema = defaultdict(int)
    for reg in registros_missoes:
        if reg['tema'] and reg['tema'] != '—':
            qtd_tema[reg['tema']] += 1
    for nome, t in temas.items():
        t['qtd_missoes'] = qtd_tema.get(nome, 0)


    return result


def extrair_ppgs(xlsx_path):
    import statistics
    from collections import defaultdict
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    TEMAS = ['TEMA 1','TEMA 2','TEMA 3','TEMA 4','TEMA 5']
    registros = []
    for r in range(3, ws.max_row + 1):
        ies = ws.cell(r,1).value
        ppg = ws.cell(r,3).value
        if not ies or not ppg or str(ies).strip().upper() == 'TOTAL': continue
        nota = ws.cell(r,4).value or 0
        doc = ws.cell(r,5).value or 0
        disc = ws.cell(r,6).value or 0
        temas_sim = []
        for i, t in enumerate(TEMAS):
            val = ws.cell(r,7+i).value
            if val and str(val).strip().upper() == 'SIM':
                temas_sim.append(f'Tema {i+1}')
        registros.append({'ies': str(ies).strip(), 'ppg': str(ppg).strip(), 'nota': nota,
                           'docentes': doc, 'discentes': disc, 'temas': temas_sim})

    total = len(registros)
    if total == 0:
        return None
    nota_media = sum(r['nota'] for r in registros) / total
    total_docentes = sum(r['docentes'] for r in registros)
    total_discentes = sum(r['discentes'] for r in registros)
    altas = [r for r in registros if r['nota'] >= 6]

    kpis = {'total': total, 'total_ies': len(set(r['ies'] for r in registros)),
            'nota_media': round(nota_media, 2), 'total_docentes': total_docentes,
            'docentes_media': round(total_docentes/total, 1), 'total_discentes': total_discentes,
            'discentes_media': round(total_discentes/total, 1), 'ppgs_alta_nota': len(altas),
            'ppgs_alta_nota_pct': round(len(altas)/total*100, 0)}

    mediana_geral = {'docentes': statistics.median(r['docentes'] for r in registros),
                      'discentes': statistics.median(r['discentes'] for r in registros)}

    por_ies_map = defaultdict(list)
    for r in registros: por_ies_map[r['ies']].append(r)
    por_ies = []
    for ies, regs in sorted(por_ies_map.items(), key=lambda x: -len(x[1])):
        por_ies.append({'ies': ies, 'qtd': len(regs), 'nota_media': round(sum(x['nota'] for x in regs)/len(regs), 2),
            'docentes': sum(x['docentes'] for x in regs), 'discentes': sum(x['discentes'] for x in regs),
            'mediana_docentes': statistics.median(x['docentes'] for x in regs),
            'mediana_discentes': statistics.median(x['discentes'] for x in regs)})

    mediana_por_tema = []
    for i in range(1,6):
        tema = f'Tema {i}'
        regs = [r for r in registros if tema in r['temas']]
        if regs:
            mediana_por_tema.append({'tema': tema, 'qtd': len(regs),
                'docentes': statistics.median(x['docentes'] for x in regs),
                'discentes': statistics.median(x['discentes'] for x in regs)})

    dist_notas = defaultdict(int)
    for r in registros: dist_notas[r['nota']] += 1
    distribuicao_notas = [{'nota': n, 'qtd': dist_notas[n]} for n in sorted(dist_notas.keys())]

    cobertura_tematica = []
    for i in range(1,6):
        tema = f'Tema {i}'
        qtd = sum(1 for r in registros if tema in r['temas'])
        cobertura_tematica.append({'tema': tema, 'qtd': qtd, 'pct': round(qtd/total*100,1)})
    cobertura_tematica.sort(key=lambda x: -x['qtd'])

    lista = sorted(
        [{'ppg': r['ppg'], 'ies': r['ies'], 'nota': r['nota'], 'docentes': r['docentes'], 'discentes': r['discentes']} for r in registros],
        key=lambda x: (-x['nota'], x['ies'], x['ppg']))

    return {'kpis': kpis, 'mediana_geral': mediana_geral, 'por_ies': por_ies,
            'mediana_por_tema': mediana_por_tema, 'distribuicao_notas': distribuicao_notas,
            'cobertura_tematica': cobertura_tematica, 'lista': lista}


def main():
    if len(sys.argv) < 2:
        print("Uso: python3 atualizar_dashboard.py <planilha_controle.xlsx> [planilha_ppgs.xlsx]")
        sys.exit(1)
    xlsx_path = sys.argv[1]
    data = extrair(xlsx_path)

    idx_path = pathlib.Path(__file__).parent / 'index.html'
    html = idx_path.read_text(encoding='utf-8')

    # Preserva os dados de PPGs já existentes no painel, a menos que uma planilha de PPGs seja informada
    m = re.search(r"const DATA = (.*?);\n", html, flags=re.S)
    ppgs_atual = None
    if m:
        try:
            ppgs_atual = json.loads(m.group(1)).get('ppgs')
        except Exception:
            ppgs_atual = None

    if len(sys.argv) >= 3:
        ppgs_novo = extrair_ppgs(sys.argv[2])
        if ppgs_novo:
            data['ppgs'] = ppgs_novo
            print("Dados de PPGs atualizados a partir de", sys.argv[2])
    elif ppgs_atual:
        data['ppgs'] = ppgs_atual

    html = re.sub(r"const DATA = .*?;\n", "const DATA = " + json.dumps(data, ensure_ascii=False) + ";\n", html, count=1, flags=re.S)
    idx_path.write_text(html, encoding='utf-8')
    print("Painel atualizado:", idx_path)
    print("Agora e so: git add . && git commit -m 'atualiza dados' && git push")


if __name__ == '__main__':
    main()
